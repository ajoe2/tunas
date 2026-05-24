"""Dev-only: convert USA Swimming motivational-standard .xlsx sheets to bundled JSON.

Reads the six B–AAAA workbooks (each with per-age-group sheets) from the source
directory and writes ``src/tunas/_data/standards-2025-2028.json``. NOT shipped in
the wheel; run manually when standards change:

    uv run python scripts/convert_standards.py [SOURCE_DIR]

``SOURCE_DIR`` defaults to the original tunas-cli data directory.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

from tunas.enums import Stroke
from tunas.event import Event

# file basename -> TimeStandard name (the six motivational standards)
_FILES: dict[str, str] = {
    "B-2028.xlsx": "B",
    "BB-2028.xlsx": "BB",
    "A-2028.xlsx": "A",
    "AA-2028.xlsx": "AA",
    "AAA-2028.xlsx": "AAA",
    "AAAA-2028.xlsx": "AAAA",
}

# Sheet order (after the "Source" sheet) -> age-group label.
_AGE_GROUPS = ["10_U", "11_12", "13_14", "15_16", "17_18"]

# Column header -> (course token, sex code).
_COLUMNS: dict[str, tuple[str, str]] = {
    "SCY-F": ("SCY", "F"),
    "LCM-F": ("LCM", "F"),
    "SCM-F": ("SCM", "F"),
    "SCY-M": ("SCY", "M"),
    "LCM-M": ("LCM", "M"),
    "SCM-M": ("SCM", "M"),
}

_STROKES: dict[str, Stroke] = {
    "FR": Stroke.FREESTYLE,
    "BK": Stroke.BACKSTROKE,
    "BR": Stroke.BREASTSTROKE,
    "FL": Stroke.BUTTERFLY,
    "IM": Stroke.INDIVIDUAL_MEDLEY,
    "FR RELAY": Stroke.FREESTYLE_RELAY,
    "IM RELAY": Stroke.MEDLEY_RELAY,
}

# Merged distance rows: token -> {course: actual distance}.
_MERGED_DISTANCE: dict[str, dict[str, int]] = {
    "400/500": {"SCY": 500, "SCM": 400, "LCM": 400},
    "800/1000": {"SCY": 1000, "SCM": 800, "LCM": 800},
    "1500/1650": {"SCY": 1650, "SCM": 1500, "LCM": 1500},
}

_DEFAULT_SOURCE = Path("/storage_slow/ajoe/code/tunas-cli/data/timeStandards")
_OUTPUT = (
    Path(__file__).resolve().parent.parent / "src" / "tunas" / "_data" / "standards-2025-2028.json"
)


def _parse_centiseconds(raw: object) -> int | None:
    if raw is None:
        return None
    text = str(raw).strip().rstrip("*").strip()
    if not text or text == "0":
        return None
    minutes = 0
    sec_part = text
    if ":" in text:
        m, sec_part = text.split(":", 1)
        minutes = int(m)
    seconds, frac = sec_part.split(".")
    return minutes * 6000 + int(seconds) * 100 + int((frac + "00")[:2])


def _resolve_event(label: str, course: str) -> Event | None:
    parts = label.split(None, 1)
    if len(parts) != 2:
        return None
    dist_token, stroke_token = parts[0], parts[1].strip().upper()
    stroke = _STROKES.get(stroke_token)
    if stroke is None:
        return None
    from tunas.enums import Course

    course_enum = {"SCY": Course.SCY, "SCM": Course.SCM, "LCM": Course.LCM}[course]
    if dist_token in _MERGED_DISTANCE:
        distance = _MERGED_DISTANCE[dist_token][course]
    elif dist_token.isdigit():
        distance = int(dist_token)
    else:
        return None
    return Event.find(distance, stroke, course_enum)


def convert(source_dir: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for filename, standard in _FILES.items():
        wb = openpyxl.load_workbook(source_dir / filename, read_only=True, data_only=True)
        sheets = wb.sheetnames[1:]  # skip "Source"
        for sheet_name, age_group in zip(sheets, _AGE_GROUPS, strict=False):
            ws = wb[sheet_name]
            data = list(ws.iter_rows(values_only=True))
            header = [str(c).strip() if c else "" for c in data[0]]
            col_index = {h: i for i, h in enumerate(header)}
            for record in data[1:]:
                label = record[0]
                if not label:
                    continue
                label = str(label).strip()
                for col, (course, sex) in _COLUMNS.items():
                    if col not in col_index:
                        continue
                    cs = _parse_centiseconds(record[col_index[col]])
                    if cs is None:
                        continue
                    event = _resolve_event(label, course)
                    if event is None:
                        continue
                    rows.append(
                        {
                            "standard": standard,
                            "age_group": age_group,
                            "sex": sex,
                            "event": event.name,
                            "cutoff_centiseconds": cs,
                        }
                    )
        wb.close()
    return rows


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else _DEFAULT_SOURCE
    rows = convert(source)
    # Detect duplicates (would break the runtime index).
    seen: set[tuple[object, ...]] = set()
    for r in rows:
        key = (r["standard"], r["age_group"], r["sex"], r["event"])
        if key in seen:
            raise SystemExit(f"duplicate standard row: {key}")
        seen.add(key)
    payload = {
        "version": "2025-2028",
        "season_start": "2024-09-01",
        "season_end": "2028-08-31",
        "source_notes": "USA Swimming 2025-2028 motivational standards (B-AAAA)",
        "standards": rows,
    }
    _OUTPUT.write_text(json.dumps(payload, indent=1) + "\n")
    print(f"wrote {len(rows)} standards to {_OUTPUT}")


if __name__ == "__main__":
    main()
